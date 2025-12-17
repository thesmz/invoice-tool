import streamlit as st
import pandas as pd
from google.cloud import documentai_v1 as documentai
from google.api_core.client_options import ClientOptions
import os
import re
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import fitz  # PyMuPDF
from PIL import Image

# --- CONFIGURATION ---
st.set_page_config(page_title="Invoice Processor V4", layout="wide", page_icon="📄")

if 'processed_data' not in st.session_state:
    st.session_state.processed_data = []
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False

# --- HELPER FUNCTIONS ---
def clean_amount(amount_str):
    """Clean string to number"""
    if not amount_str: return 0.0
    cleaned = re.sub(r'[^\d.]', '', str(amount_str))
    try:
        val = float(cleaned)
        return int(val) if val.is_integer() else val
    except:
        return 0.0

def get_pdf_image(file_bytes, page_num=0):
    """Render a specific page of a PDF as an image for preview"""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        # Safety check for page count
        if page_num >= len(doc): page_num = 0
        if page_num < 0: page_num = 0
        
        page = doc.load_page(page_num)
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2)) # 2x zoom for clarity
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return img
    except Exception as e:
        return None

def get_page_count(file_bytes):
    """Get total pages in PDF"""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        return len(doc)
    except:
        return 1

def extract_vendor_and_items(filename):
    """Regex logic to guess Vendor and Items"""
    vendor = ""
    items = ""
    name_without_ext = os.path.splitext(filename)[0]
    
    items_match = re.search(r'[（(]([^）)]+)[）)]', name_without_ext)
    if items_match:
        items = items_match.group(1).strip()
        name_for_vendor = re.sub(r'[（(][^）)]+[）)]', '', name_without_ext)
    else:
        name_for_vendor = name_without_ext
        
    name_for_vendor = re.sub(r'^[〇○◯]?\d{6}\s*[-－]\s*', '', name_for_vendor)
    name_for_vendor = re.sub(r'(未払金|買掛金)(計上済|補助なし計上済|[(（]補助なし[)）]計上済)?', '', name_for_vendor)
    vendor = name_for_vendor.strip().strip('-_. ')
    
    if not vendor: vendor = name_without_ext
    return vendor, items

def process_document_ai(file_content, mime_type, project_id, loc, proc_id, credentials_file):
    with open("temp_creds.json", "wb") as f:
        f.write(credentials_file.getbuffer())
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = "temp_creds.json"

    opts = ClientOptions(api_endpoint=f"{loc}-documentai.googleapis.com")
    client = documentai.DocumentProcessorServiceClient(client_options=opts)
    name = client.processor_path(project_id, loc, proc_id)
    
    raw_document = documentai.RawDocument(content=file_content, mime_type=mime_type)
    request = documentai.ProcessRequest(name=name, raw_document=raw_document)
    result = client.process_document(request=request)
    return result.document

def append_to_excel(existing_file, new_data):
    """Append new data to an existing Excel file"""
    wb = openpyxl.load_workbook(existing_file)
    
    if "Invoice Summary" in wb.sheetnames:
        ws = wb["Invoice Summary"]
    else:
        ws = wb.create_sheet("Invoice Summary")
    
    next_row = ws.max_row + 1
    
    for item in new_data:
        fb_total = 0
        others_total = 0
        for amt in item['amounts']:
            val = float(amt['value'])
            cat = amt['category']
            if cat == 'FB Amount': fb_total += val
            elif cat == 'Others': others_total += val
            elif cat == 'Divide (50/50)':
                fb_total += val / 2
                others_total += val / 2
        
        if fb_total > 0 or others_total > 0:
            ws.cell(row=next_row, column=1, value=item['vendor_name'])
            ws.cell(row=next_row, column=2, value=item['items_desc'])
            ws.cell(row=next_row, column=3, value=fb_total if fb_total > 0 else '')
            ws.cell(row=next_row, column=4, value=others_total if others_total > 0 else '')
            
            ws.cell(row=next_row, column=3).number_format = '#,##0'
            ws.cell(row=next_row, column=4).number_format = '#,##0'
            next_row += 1
            
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def create_new_excel(data_list):
    """Create fresh Excel file"""
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws = wb.active
    ws.title = "Invoice Summary"
    headers = ['Vendor Name', 'Items', 'FB Amount (Tax incld.)', 'Others Amount (Tax incld.)']
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for item in data_list:
        fb_total = 0
        others_total = 0
        for amt in item['amounts']:
            val = float(amt['value'])
            cat = amt['category']
            if cat == 'FB Amount': fb_total += val
            elif cat == 'Others': others_total += val
            elif cat == 'Divide (50/50)':
                fb_total += val / 2
                others_total += val / 2
        
        ws.append([
            item['vendor_name'],
            item['items_desc'],
            fb_total if fb_total > 0 else '',
            others_total if others_total > 0 else ''
        ])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    for row in range(2, ws.max_row + 1):
        for col in [3, 4]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- SIDEBAR ---
with st.sidebar:
    st.header("⚙️ Configuration")
    creds_file = st.file_uploader("Credentials JSON", type="json")
    with st.expander("Advanced Settings"):
        project_id = st.text_input("Project ID", value="receipt-processor-479605")
        location = st.selectbox("Location", ["us", "eu"], index=0)
        processor_id = st.text_input("Processor ID", value="88cff36a297265dc")

# --- MAIN APP ---
st.title("📄 Invoice Processor V4")

uploaded_files = st.file_uploader("1. Upload New Invoices", type=['pdf'], accept_multiple_files=True)
start_btn = st.button("🚀 Process Batch", type="primary", disabled=(not uploaded_files or not creds_file))

if start_btn:
    st.session_state.processed_data = [] 
    progress_bar = st.progress(0)
    
    for idx, f in enumerate(uploaded_files):
        # Read file bytes once
        file_bytes = f.read()
        f.seek(0)
        
        # Calculate total pages immediately
        total_pages = get_page_count(file_bytes)
        f.seek(0)
        
        vendor, items = extract_vendor_and_items(f.name)
        
        try:
            doc = process_document_ai(file_bytes, f.type, project_id, location, processor_id, creds_file)
            
            extracted_amounts = []
            totals_by_page = {}
            for entity in doc.entities:
                if entity.type_ == 'total_amount':
                    page = entity.page_anchor.page_refs[0].page if entity.page_anchor.page_refs else 0
                    if page not in totals_by_page: totals_by_page[page] = []
                    totals_by_page[page].append(clean_amount(entity.mention_text))
            
            for page, amounts in totals_by_page.items():
                if amounts:
                    extracted_amounts.append({
                        "page": page + 1,
                        "value": max(amounts),
                        "category": "FB Amount"
                    })
            
            st.session_state.processed_data.append({
                "filename": f.name,
                "file_bytes": file_bytes,
                "page_count": total_pages, # Save total pages
                "vendor_name": vendor,
                "items_desc": items,
                "amounts": extracted_amounts
            })
            
        except Exception as e:
            st.error(f"Error {f.name}: {e}")
        progress_bar.progress((idx + 1) / len(uploaded_files))
    
    st.session_state.processing_complete = True

# --- REVIEW SECTION ---
if st.session_state.processing_complete:
    st.divider()
    st.subheader("2. Review & Edit")
    
    for i, invoice in enumerate(st.session_state.processed_data):
        with st.expander(f"📄 {invoice['vendor_name']}", expanded=(i==0)):
            col_preview, col_data = st.columns([1, 1.2])
            
            # --- LEFT COLUMN: PREVIEW ---
            with col_preview:
                st.markdown("**Invoice Preview**")
                
                # Dynamic Page Selector
                total_pgs = invoice.get('page_count', 1)
                
                # Determine default page (if an amount was found, start there, otherwise page 1)
                default_page = 1
                if invoice['amounts']:
                    default_page = invoice['amounts'][0]['page']
                if default_page > total_pgs: default_page = 1
                
                # Page Selector Widget
                selected_page = st.number_input(
                    f"Showing Page (Total {total_pgs})", 
                    min_value=1, 
                    max_value=total_pgs, 
                    value=default_page,
                    key=f"pg_sel_{i}"
                )
                
                # Render Image (0-indexed for fitz, so subtract 1)
                img = get_pdf_image(invoice['file_bytes'], selected_page - 1)
                if img:
                    st.image(img, use_container_width=True)
                else:
                    st.error("Cannot load preview")

            # --- RIGHT COLUMN: DATA ---
            with col_data:
                new_vendor = st.text_input("Vendor", value=invoice['vendor_name'], key=f"v_{i}")
                new_items = st.text_area("Items", value=invoice['items_desc'], height=1, key=f"d_{i}")
                
                st.session_state.processed_data[i]['vendor_name'] = new_vendor
                st.session_state.processed_data[i]['items_desc'] = new_items
                
                st.markdown("---")
                st.markdown("**Amounts**")
                
                amounts_to_rem = []
                for j, amount in enumerate(invoice['amounts']):
                    c1, c2, c3 = st.columns([2, 3, 1])
                    
                    new_val = c1.number_input(f"Page {amount['page']} Amount", value=float(amount['value']), key=f"val_{i}_{j}")
                    st.session_state.processed_data[i]['amounts'][j]['value'] = new_val
                    
                    cat_opts = ["FB Amount", "Others", "Divide (50/50)", "None"]
                    curr_idx = cat_opts.index(amount['category']) if amount['category'] in cat_opts else 0
                    new_cat = c2.selectbox("Category", cat_opts, index=curr_idx, key=f"cat_{i}_{j}", label_visibility="collapsed")
                    st.session_state.processed_data[i]['amounts'][j]['category'] = new_cat
                    
                    if c3.button("🗑️", key=f"del_{i}_{j}"): amounts_to_rem.append(j)
                
                if amounts_to_rem:
                    for idx in sorted(amounts_to_rem, reverse=True):
                        del st.session_state.processed_data[i]['amounts'][idx]
                    st.rerun()
                
                if st.button("➕ Add Manual Amount", key=f"add_{i}"):
                    # Default new amount to the currently viewed page in preview
                    st.session_state.processed_data[i]['amounts'].append({
                        "page": selected_page, 
                        "value": 0.0, 
                        "category": "FB Amount"
                    })
                    st.rerun()

    # --- EXPORT SECTION ---
    st.divider()
    st.subheader("3. Export")
    
    export_mode = st.radio("Choose Export Mode:", ["Create New Excel File", "Append to Existing Excel"])
    
    final_data = None
    file_name = f"invoice_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    if export_mode == "Create New Excel File":
        if st.button("💾 Download New File"):
            final_data = create_new_excel(st.session_state.processed_data)
            
    else: # Append Mode
        existing_excel = st.file_uploader("Upload your master Excel file (e.g. October.xlsx)", type=['xlsx'])
        if existing_excel:
            if st.button("💾 Merge & Download"):
                try:
                    final_data = append_to_excel(existing_excel, st.session_state.processed_data)
                    file_name = existing_excel.name
                except Exception as e:
                    st.error(f"Error merging excel: {e}")

    if final_data:
        st.download_button(
            label="📥 Click to Download Result",
            data=final_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
